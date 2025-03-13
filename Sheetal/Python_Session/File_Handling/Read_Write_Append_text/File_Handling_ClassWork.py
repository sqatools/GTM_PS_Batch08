"""
# Mode to access the file
# read mode (r) : we can open file in read mode to read its content.

# write mode (w) : -> we can open file in write mode if we want to update the content of the file.
                   -> When we file in write mode and add some content to the file, then previous content
                      will overwrite in new content.

# append mode (a) : -> We can open file in append mode to update the content of the file, and
                       append mode add new content to the file at end of the file.
                    ->  Append mode keep the original data as well as updated content.

# Note :  function to open file:
file = open(filepath, mode)
"""


# read content from file
def read_file_content(file_path):
    file_obj = open(file_path, "r")
    data = file_obj.read()
    print("is file readable :", file_obj.readable())
    print("file name :", file_obj.name)
    print("is file closed :", file_obj.closed)
    print(data)
    file_obj.close()
    print("is file closed :", file_obj.closed)
    # print(file_obj.read())


# read file from current location
# read_file_content("read_data.txt")

# read file from different location
# read_file_content("E:\\Filesdata\\count_name.txt")

"""
# Mode to access the file
# read mode (r) : we can open file in read mode to read its content.

# write mode (w) : -> we can open file in write mode if we want to update the content of the file.
                   -> When we file in write mode and add some content to the file, then previous content
                      will overwrite in new content.

# append mode (a) : -> We can open file in append mode to update the content of the file, and 
                       append mode add new content to the file at end of the file.
                    ->  Append mode keep the original data as well as updated content.

# Note :  function to open file:
file = open(filepath, mode)
"""


# read content from file
def read_file_content(file_path):
    file_obj = open(file_path, "r")
    data = file_obj.read()
    print("is file readable :", file_obj.readable())
    print("file name :", file_obj.name)
    print("is file closed :", file_obj.closed)
    print(data)
    file_obj.close()
    print("is file closed :", file_obj.closed)
    # print(file_obj.read())


# read file from current location
# read_file_content("read_data.txt")

# read file from different location
# read_file_content("E:\\Filesdata\\count_name.txt")


# write content to the file with write mode.

def write_content_to_file(file_path, new_content):
    file_obj = open(file_path, "w")
    file_obj.write(new_content)
    file_obj.close()


# 1. write content to existing file
# write_content_to_file("new_file.txt", "Hello Good Morning")

# 2. write content to non-existing file : If file is not available then it will create new file
#    and add content to the file.
# write_content_to_file("new_file2.txt", "Very good evening")


# Append content to the file

def append_content_to_file(filepath, content):
    file = open(filepath, "a")
    file.write(content)
    file.close()


# 1. Add content to existing file: It will add content at end of existing file
# append_content_to_file("append_data.txt", "Adding new content to file\n")

# 1. Add content to non-existing file: It will create new file and add content to the file
# append_content_to_file("append_data_new.txt", "Writing first line\n")
para = """
Nathan Ellis to Hardik Pandya, 
out Caught by Maxwell!! Hardik Pandya c 
Maxwell b Nathan Ellis 28(24) [4s-1 6s-3]
Nathan Ellis to Hardik Pandya, THATS OUT!! Caught!!
India are just one hit away!
47.4
Nathan Ellis to Hardik Pandya, 
no run, wide yorker, Hardik Pandya 
reaches out for it and gets an under-edge that rolls towards the keeper
47.3
Nathan Ellis to Hardik Pandya, 
FOUR, with mid-off and mid-on inside, 
Hardik Pandya goes straight even though the length is
not on the fuller side. Big backlift, clears the front 
leg and hammers this short of a good length ball over the 
bowler's head
"""
# append_content_to_file("append_data_new.txt", para)


print("_" * 50)


#######################################################
# context manager :  context manager has it own inbuilt enter and exist method, when ever we open
#                     any file with context manager then no need to close it explicitly.

def read_file_content_with_context(filepath):
    with open(filepath, "r") as file_obj:
        data = file_obj.read()
        print(data)
        print("is file closed in context manager :", file_obj.closed)  # False

    print("is file closed outside of context manager :", file_obj.closed)  # True


# read_file_content_with_context("read_data.txt")


print("_" * 50)


#######################################################
# File read options.
# 1. read specific number of bytes
# 2. read number of lines
# 3. read list of lines


# 1. read specific number of bytes/chars
def read_number_of_bytes(filepath, num_of_bytes):
    with open(filepath, "r") as file_obj:
        data = file_obj.read(num_of_bytes)
        print(data)


read_number_of_bytes("read_data.txt", 10)
# 1.good mor

read_number_of_bytes("read_data.txt", 50)
"""
1.good morning
2.Learning Python
3.Hope you are do
"""

print("_" * 50)


#############################
# 2. read number of lines with readline() method.
def read_number_lines(filepath, num_lines):
    with open(filepath, "r") as file_obj:
        # print(file_obj.readline())
        # print(file_obj.readline())
        # print(file_obj.readline())
        for _ in range(num_lines):
            print(file_obj.readline(), end="")


# read_number_lines("read_data.txt", 15)


print("#" * 50)


############################################################
# 3. read list of lines with readlines() method:  readlines method return list of lines from given file

def read_list_of_lines(filepath):
    with open(filepath, "r") as file:
        lines_list = file.readlines()
        print(lines_list)
        return lines_list


# read_list_of_lines("read_data.txt")


def read_specific_number_lines(filepath, start, stop):
    lines_list = read_list_of_lines(filepath)
    for i in range(start - 1, stop):
        print(lines_list[i], end="")


read_specific_number_lines("read_data.txt", 5, 10)

"""
5.Python is easy to learn.
6.Hope you are doing good.
7.India is un beatable in CT.
8.Python is easy to learn.
9.Hope you are doing good.
10.India is un beatable in CT.
"""

# write content to the file with write mode.

def write_content_to_file(file_path, new_content):
    file_obj = open(file_path, "w")
    file_obj.write(new_content)
    file_obj.close()


write_content_to_file("new_file.txt", "Hello Good Morning")


# read all the email and phone number given file.

def read_email_phone(file_path):
    """
    This function will read file and return the list of emails and
    phone numbers.
    :param file_path:
    :return:
    """
    email_list = []
    phone_list = []
    # open file
    with open(file_path, "r") as file:
        # read file data
        data = file.read()

    # split file data in word list
    word_list = data.split()

    # Apply loop on word list
    for word in word_list:
        # check if word contains @, then consider as email
        if "@" in word:
            # append email to email_list
            email_list.append(word)
        # check if word len is 10, and it contains only digit, then consider as phone number.
        elif len(word) == 10 and word.isdigit():
            # append phone number to phone_list
            phone_list.append(word)

    return email_list, phone_list


# emails, phones = read_email_phone("read_data.txt")
# print("emails :", emails)
# # emails : ['user2@yahoo.com', 'user1@gmail.com', 'user2@facebook.com']
# print("phone numbers :", phones)
# phone numbers : ['8845465644', '5445465644', '9954656447']



# Q2 : replace the specific word in file
def replace_word_from_file(filepath, word1, word2):
    with open(filepath, "r") as file:
        file_data = file.read()

    updated_content = file_data.replace(word1, word2)

    with open(filepath, "w") as file:
        file.write(updated_content)

replace_word_from_file("read_data.txt", "Python", "JAVA")


#Q3 : Read content from file1 and file2 and add content to the file3.



"""
command to install library
 -> pip install openpyxl
# command to check list of installed python packages
 -> pip list

"""
import openpyxl


def read_excel_file(file_path, sheet_name, cell_name):
    wb = openpyxl.load_workbook(file_path)
    # print(wb.__dict__)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


# read_excel_file("test_data.xlsx", sheet_name='Sheet1', cell_name="A3")
# read_excel_file("test_data.xlsx", sheet_name='Sheet2', cell_name="C1")
# read_excel_file("test_data.xlsx", "Sheet3", "D1")
# read_excel_file("test_data.xlsx", "Kartik", "A13")


# Read random cells values from excel sheet
"""
randon_list = [4, 6, 9, 1, 5]
for val in randon_list:
    read_excel_file("test_data.xlsx", sheet_name='Sheet1', cell_name=f"A{val}")
"""


# read specific number of cell values in given range.
def read_excel_file_specific_cells(file_path, sheet_name, cell_name, start, end):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    for i in range(start, end + 1):
        cell = sheet[f"{cell_name}{i}"]
        print(cell.value)


#read_excel_file_specific_cells("test_data.xlsx", "Sheet1", "A", 3, 7)


def read_excel_file2(file_path, sheet_name, cell_name):
    cell = openpyxl.load_workbook(file_path)[sheet_name][cell_name]
    print(cell.value)

#read_excel_file2("test_data.xlsx", "Sheet1", "A1")

# read all cell values without range
def read_excel_file_with_max_rows(file_path, sheet_name, cell_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    max_row = sheet.max_row # it will return total active rows
    print(max_row) #
    for i in range(1, max_row+1):
        cell = sheet[f"{cell_name}{i}"]
        print(cell.value)

read_excel_file_with_max_rows("test_data.xlsx", "Sheet1", "A")


# read all cell values from all the columns
def read_excel_file_with_max_rows_coln(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    max_row = sheet.max_row # it will return total active rows
    print(max_row) #
    max_colum = sheet.max_column
    # i = row, j = col
    for i in range(1, max_row+1):
        for j in range(1, max_colum+1):
            cell = sheet.cell(row=i, column=j)
            print(cell.value, end=" ")
        print()

read_excel_file_with_max_rows_coln("test_data.xlsx", "Sheet1")
