"""
command to install library
 -> pip install openpyxl
# command to check list of installed python packages
 -> pip list

"""
import openpyxl


def read_excel_file(file_path, sheet_name, cell_name):
    wb = openpyxl.load_workbook(file_path)
    # print(wb.__dict__)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


# read_excel_file("test_data.xlsx", sheet_name='Sheet1', cell_name="A3")
# read_excel_file("test_data.xlsx", sheet_name='Sheet2', cell_name="C1")
# read_excel_file("test_data.xlsx", "Sheet3", "D1")
# read_excel_file("test_data.xlsx", "Kartik", "A13")


# Read random cells values from excel sheet
"""
randon_list = [4, 6, 9, 1, 5]
for val in randon_list:
    read_excel_file("test_data.xlsx", sheet_name='Sheet1', cell_name=f"A{val}")
"""


# read specific number of cell values in given range.
def read_excel_file_specific_cells(file_path, sheet_name, cell_name, start, end):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    for i in range(start, end + 1):
        cell = sheet[f"{cell_name}{i}"]
        print(cell.value)


#read_excel_file_specific_cells("test_data.xlsx", "Sheet1", "A", 3, 7)


def read_excel_file2(file_path, sheet_name, cell_name):
    cell = openpyxl.load_workbook(file_path)[sheet_name][cell_name]
    print(cell.value)

#read_excel_file2("test_data.xlsx", "Sheet1", "A1")

# read all cell values without range
def read_excel_file_with_max_rows(file_path, sheet_name, cell_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    max_row = sheet.max_row # it will return total active rows
    print(max_row) #
    for i in range(1, max_row+1):
        cell = sheet[f"{cell_name}{i}"]
        print(cell.value)

read_excel_file_with_max_rows("test_data.xlsx", "Sheet1", "A")


# read all cell values from all the columns
def read_excel_file_with_max_rows_coln(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    max_row = sheet.max_row # it will return total active rows
    print(max_row) #
    max_colum = sheet.max_column
    # i = row, j = col
    for i in range(1, max_row+1):
        for j in range(1, max_colum+1):
            cell = sheet.cell(row=i, column=j)
            print(cell.value, end=" ")
        print()

read_excel_file_with_max_rows_coln("test_data.xlsx", "Sheet1")
